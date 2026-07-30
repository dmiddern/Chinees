import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORDS_PATHS = [
    ROOT / "src" / "data" / "words.json",
    ROOT / "src" / "data" / "words-hsk4.json",
    ROOT / "src" / "data" / "words-hsk5.json",
    ROOT / "src" / "data" / "words-hsk6.json",
]
SOURCE_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "scripts" / "data" / "HSK_Level_7-9.xlsx"
EXPECTED_COUNT = 5605

WORD_TYPES = {
    "名": "noun",
    "动": "verb",
    "形": "adjective",
    "副": "adverb",
    "代": "pronoun",
    "介": "preposition",
    "连": "conjunction",
    "叹": "interjection",
    "助": "auxiliary",
    "量": "classifier",
    "数": "numeral",
}


def translate_word_type(value):
    if not value:
        return ""
    parts = str(value).replace("（", "").replace("）", "").replace("(", "").replace(")", "")
    return "/".join(WORD_TYPES.get(part.strip(), part.strip()) for part in parts.split("、") if part.strip())


base_words = []
for words_path in WORDS_PATHS:
    with words_path.open(encoding="utf-8") as handle:
        base_words.extend(json.load(handle))
sheet = load_workbook(SOURCE_PATH, read_only=True, data_only=True).active
rows = list(sheet.iter_rows(min_row=2, values_only=True))

if len(rows) != EXPECTED_COUNT:
    raise RuntimeError(f"HSK 7-9: found {len(rows)} rows, expected {EXPECTED_COUNT}.")

next_id = max(word["id"] for word in base_words)
advanced_words = []
for simplified, _traditional, pinyin, _zhuyin, _level, word_type, _frequency, definition in rows:
    next_id += 1
    advanced_words.append(
        {
            "id": next_id,
            "level": "7-9",
            "hanzi": str(simplified).strip(),
            "pinyin": str(pinyin).strip(),
            "wordType": translate_word_type(word_type),
            "meaningNl": str(definition).strip(),
            "meaningLanguage": "en",
            "example": "",
            "notes": "",
            "source": "https://github.com/krmanik/HSK-3.0/tree/main/New%20HSK%20(2025)",
        }
    )

chunk_size = (len(advanced_words) + 3) // 4
for index in range(4):
    output_path = ROOT / "src" / "data" / f"words-hsk7-9-{index + 1}.json"
    with output_path.open("w", encoding="utf-8") as handle:
        json.dump(
            advanced_words[index * chunk_size : (index + 1) * chunk_size],
            handle,
            ensure_ascii=False,
            indent=2,
        )
        handle.write("\n")

print(f"Imported {len(advanced_words)} HSK 7-9 entries; {len(base_words) + len(advanced_words)} total.")
